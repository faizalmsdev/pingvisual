import openpyxl
import requests
import json
import time
from typing import Optional

class SimpleJobCreator:
    def __init__(self, base_url: str = "http://localhost:5000"):
        self.base_url = base_url
        self.session = requests.Session()
        self.token = None
        self.user_id = None
        
    def login(self, email: str, password: str) -> bool:
        """Login to the API and get authentication token"""
        try:
            login_url = f"{self.base_url}/api/auth/login"
            login_data = {
                "email": email,
                "password": password
            }
            
            print(f"🔐 Logging in with email: {email}")
            response = self.session.post(login_url, json=login_data)
            
            if response.status_code == 200:
                data = response.json()
                self.token = data['token']
                self.user_id = data['user']['user_id']
                print(f"✅ Login successful!")
                print(f"   User ID: {self.user_id}")
                print(f"   Email: {data['user']['email']}")
                return True
            else:
                print(f"❌ Login failed: {response.json().get('error', 'Unknown error')}")
                return False
                
        except Exception as e:
            print(f"❌ Login error: {e}")
            return False
    
    def create_job(self, name: str, url: str, check_interval_minutes: int) -> Optional[str]:
        """Create a monitoring job"""
        try:
            jobs_url = f"{self.base_url}/api/jobs"
            job_data = {
                "name": name,
                "url": url,
                "check_interval_minutes": check_interval_minutes
            }
            
            # Add authentication header
            headers = {
                "Authorization": f"Bearer {self.token}",
                "Content-Type": "application/json"
            }
            
            print(f"📝 Creating job: {name}")
            print(f"   URL: {url}")
            print(f"   Interval: {check_interval_minutes} minutes ({check_interval_minutes/60:.1f} hours)")
            
            response = self.session.post(jobs_url, json=job_data, headers=headers)
            
            if response.status_code == 201:
                data = response.json()
                job_id = data['job_id']
                print(f"✅ Job created successfully!")
                print(f"   Job ID: {job_id}")
                return job_id
            else:
                error_msg = response.json().get('error', 'Unknown error')
                print(f"❌ Failed to create job: {error_msg}")
                return None
                
        except Exception as e:
            print(f"❌ Error creating job: {e}")
            return None
    
    def load_jobs_from_excel(self, excel_file: str) -> list:
        """Load job data from Excel file using openpyxl"""
        try:
            print(f"📊 Reading Excel file: {excel_file}")
            
            # Load workbook
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            print(f"📋 Excel file loaded successfully!")
            print(f"   Sheet name: {sheet.title}")
            print(f"   Total rows: {sheet.max_row}")
            print(f"   Total columns: {sheet.max_column}")
            
            # Read headers (first row)
            headers = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value else f"Column_{col}")
            
            print(f"   Headers: {headers}")
            
            # Extract job data (skip header row, start from row 2)
            jobs = []
            for row_num in range(2, sheet.max_row + 1):
                # Get URL from first column (column A)
                url_cell = sheet.cell(row=row_num, column=1).value
                url = str(url_cell).strip() if url_cell else ""
                
                # Get job name from second column (column B)
                name_cell = sheet.cell(row=row_num, column=2).value
                job_name = str(name_cell).strip() if name_cell else f"Job Row {row_num}"
                
                # Skip empty URLs
                if not url or url.lower() in ['none', 'nan', '']:
                    print(f"⚠️  Skipping row {row_num}: Empty URL")
                    continue
                
                # Add http:// if no protocol specified
                if not url.startswith(('http://', 'https://')):
                    url = 'https://' + url
                
                jobs.append({
                    'name': job_name,
                    'url': url,
                    'row': row_num
                })
            
            workbook.close()
            print(f"\n✅ Loaded {len(jobs)} valid jobs from Excel file")
            
            # Show preview of jobs
            print("\n📑 Jobs to be created:")
            for i, job in enumerate(jobs[:5], 1):  # Show first 5
                print(f"   {i}. {job['name']} - {job['url']}")
            if len(jobs) > 5:
                print(f"   ... and {len(jobs) - 5} more jobs")
            
            return jobs
            
        except Exception as e:
            print(f"❌ Error reading Excel file: {e}")
            return []
    
    def create_jobs_from_excel(self, excel_file: str, check_interval_hours: int = 8):
        """Create jobs from Excel file"""
        # Convert hours to minutes
        check_interval_minutes = check_interval_hours * 60
        
        print(f"🚀 Starting job creation process")
        print(f"   Excel file: {excel_file}")
        print(f"   Check interval: {check_interval_hours} hours ({check_interval_minutes} minutes)")
        print("=" * 60)
        
        # Load jobs from Excel
        jobs = self.load_jobs_from_excel(excel_file)
        
        if not jobs:
            print("❌ No jobs to create")
            return
        
        # Confirm before creating
        print(f"\n❓ Ready to create {len(jobs)} jobs. Continue? (y/n): ", end="")
        confirm = input().lower().strip()
        if confirm != 'y':
            print("❌ Operation cancelled")
            return
        
        # Create each job
        created_jobs = []
        failed_jobs = []
        
        for i, job in enumerate(jobs, 1):
            print(f"\n🔄 Processing job {i}/{len(jobs)} (Row {job['row']})")
            
            job_id = self.create_job(
                name=job['name'],
                url=job['url'],
                check_interval_minutes=check_interval_minutes
            )
            
            if job_id:
                created_jobs.append({**job, 'job_id': job_id})
            else:
                failed_jobs.append(job)
            
            # Small delay between requests
            time.sleep(1)
        
        # Summary
        print("\n" + "=" * 60)
        print("📊 JOB CREATION SUMMARY")
        print("=" * 60)
        print(f"✅ Successfully created: {len(created_jobs)} jobs")
        print(f"❌ Failed to create: {len(failed_jobs)} jobs")
        
        if created_jobs:
            print("\n✅ CREATED JOBS:")
            for job in created_jobs:
                print(f"   • {job['name']} - {job['url']} (ID: {job['job_id']})")
        
        if failed_jobs:
            print("\n❌ FAILED JOBS:")
            for job in failed_jobs:
                print(f"   • {job['name']} - {job['url']} (Row: {job['row']})")
        
        print(f"\n🎯 Total jobs processed: {len(jobs)}")
        if jobs:
            print(f"🎉 Success rate: {len(created_jobs)/len(jobs)*100:.1f}%")

def main():
    """Main function to run the job creator"""
    print("🔧 Web Change Monitor - Bulk Job Creator (Simple Version)")
    print("=" * 60)
    
    # Configuration
    EMAIL = "faizalmohamed.vi@gmail.com"
    PASSWORD = "Faizal@123"
    EXCEL_FILE = "job-settings.xlsx"
    CHECK_INTERVAL_HOURS = 8  # Check every 8 hours
    
    # Create job creator instance
    creator = SimpleJobCreator()
    
    # Login
    if not creator.login(EMAIL, PASSWORD):
        print("❌ Cannot proceed without login. Exiting...")
        return
    
    # Create jobs from Excel
    creator.create_jobs_from_excel(EXCEL_FILE, CHECK_INTERVAL_HOURS)

if __name__ == "__main__":
    main()
